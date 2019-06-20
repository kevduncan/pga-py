import bs4, requests
from openpyxl import Workbook, load_workbook

# urls for strokes gained: total, off the tee, approach, around the green, putting
sg_urls = ['https://www.pgatour.com/stats/stat.02675.html',
'https://www.pgatour.com/stats/stat.02567.html',
'https://www.pgatour.com/stats/stat.02568.html',
'https://www.pgatour.com/stats/stat.02569.html',
'https://www.pgatour.com/stats/stat.02564.html'
]

wb = Workbook()
wb = load_workbook('pga.xlsx')
sheets = wb.sheetnames
sheet_num = 0

for url in sg_urls:
    sheet_name = sheets[sheet_num]
    sheet = wb[sheet_name]
    sheet.delete_cols(0,12)

    res = requests.get(url)
    res.raise_for_status()
    urlSoup = bs4.BeautifulSoup(res.text, 'html.parser')
    table = urlSoup.find('table', {'id': 'statsTable'})

    table_head = table.find('thead')
    headers = table_head.find_all('th')
    header_vals = [ele.text.strip() for ele in headers]
    table_body = table.find('tbody')
    rows = table_body.find_all('tr')

    for i in range(len(header_vals)):
        col = i + 1
        sheet.cell(1, col, header_vals[i])

    for i in range(len(rows)):
        row_num = i + 2
        row_data = rows[i]

        cols = row_data.find_all('td')
        cols = [ele.text.strip() for ele in cols]

        for j in range(len(cols)):
            col_num = j + 1
            col_data = cols[j]
            try:
                sheet.cell(row_num, col_num, float(col_data))
            except ValueError:
                sheet.cell(row_num, col_num, col_data)
    
    sheet_num += 1

wb.save('pga.xlsx')
        

    